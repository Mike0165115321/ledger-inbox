"""
Export Service — CSV and Excel (openpyxl) export of transactions.

Supports:
- All transactions
- Per project
- Per year (tax summary)
"""

import csv
import io
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session

from ..db.models import Transaction, Project


def export_csv(
    db: Session,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
) -> str:
    """Export transactions as CSV string."""
    query = db.query(Transaction)

    if project_id:
        query = query.filter(Transaction.project_id == project_id)
    if year:
        query = query.filter(Transaction.transaction_datetime.like(f"{year}%"))

    txs = query.order_by(Transaction.transaction_datetime.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "วันที่", "ประเภท", "หมวดหมู่", "จำนวนเงิน", "สกุลเงิน",
        "ผู้ส่ง", "ผู้รับ", "ธนาคาร", "เลขอ้างอิง", "หมายเหตุ",
        "Confidence", "Review Status", "Duplicate Status", "Project ID",
    ])

    for tx in txs:
        writer.writerow([
            tx.transaction_datetime.isoformat() if tx.transaction_datetime else "",
            tx.type,
            tx.category or "",
            tx.amount,
            tx.currency,
            tx.sender_name or "",
            tx.receiver_name or "",
            tx.bank_or_wallet or "",
            tx.reference_no or "",
            tx.note or "",
            tx.confidence,
            tx.review_status,
            tx.duplicate_status,
            tx.project_id or "",
        ])

    return output.getvalue()


def export_excel(
    db: Session,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
) -> bytes:
    """Export transactions as Excel (.xlsx) bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    query = db.query(Transaction)

    if project_id:
        query = query.filter(Transaction.project_id == project_id)
    if year:
        query = query.filter(Transaction.transaction_datetime.like(f"{year}%"))

    txs = query.order_by(Transaction.transaction_datetime.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # Currency format
    thb_format = "#,##0.00"

    headers = [
        "วันที่", "ประเภท", "หมวดหมู่", "จำนวนเงิน", "สกุลเงิน",
        "ผู้ส่ง", "ผู้รับ", "ธนาคาร", "เลขอ้างอิง", "หมายเหตุ",
        "Confidence", "Review", "Duplicate",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row, tx in enumerate(txs, 2):
        ws.cell(row=row, column=1, value=tx.transaction_datetime.isoformat() if tx.transaction_datetime else "")
        ws.cell(row=row, column=2, value=tx.type)
        ws.cell(row=row, column=3, value=tx.category or "")
        cell = ws.cell(row=row, column=4, value=tx.amount)
        cell.number_format = thb_format
        ws.cell(row=row, column=5, value=tx.currency)
        ws.cell(row=row, column=6, value=tx.sender_name or "")
        ws.cell(row=row, column=7, value=tx.receiver_name or "")
        ws.cell(row=row, column=8, value=tx.bank_or_wallet or "")
        ws.cell(row=row, column=9, value=tx.reference_no or "")
        ws.cell(row=row, column=10, value=tx.note or "")
        ws.cell(row=row, column=11, value=tx.confidence)
        ws.cell(row=row, column=12, value=tx.review_status)
        ws.cell(row=row, column=13, value=tx.duplicate_status)

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Wider for note column
    ws.column_dimensions[get_column_letter(10)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_tax_summary_excel(
    db: Session,
    year: int,
) -> bytes:
    """Export tax year summary as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    txs = (
        db.query(Transaction)
        .filter(Transaction.transaction_datetime.like(f"{year}%"))
        .all()
    )

    income_txs = [t for t in txs if t.type == "income"]
    expense_txs = [t for t in txs if t.type == "expense"]

    total_income = sum(t.amount for t in income_txs)
    total_expense = sum(t.amount for t in expense_txs)
    profit = total_income - total_expense

    wb = Workbook()
    ws = wb.active
    ws.title = f"Tax Summary {year}"

    # Title
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value=f"สรุปรายรับ-รายจ่าย ปี {year}").font = Font(bold=True, size=14)

    # Summary
    headers = ["รายการ", "จำนวนเงิน (บาท)", "จำนวนรายการ"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h).font = Font(bold=True)

    ws.cell(row=4, column=1, value="รายรับรวม")
    ws.cell(row=4, column=2, value=total_income)
    ws.cell(row=4, column=3, value=len(income_txs))

    ws.cell(row=5, column=1, value="รายจ่ายรวม")
    ws.cell(row=5, column=2, value=total_expense)
    ws.cell(row=5, column=3, value=len(expense_txs))

    ws.cell(row=6, column=1, value="กำไร (ประมาณการ)")
    ws.cell(row=6, column=2, value=profit)
    ws.cell(row=6, column=1).font = Font(bold=True)
    ws.cell(row=6, column=2).font = Font(bold=True, color="006100" if profit >= 0 else "9C0006")

    # Category breakdown
    ws.cell(row=8, column=1, value="แยกตามหมวดหมู่ (รายจ่าย)").font = Font(bold=True)

    cat_map: dict[str, float] = {}
    for t in expense_txs:
        cat = t.category or "ไม่ระบุ"
        cat_map[cat] = cat_map.get(cat, 0) + t.amount

    ws.cell(row=9, column=1, value="หมวดหมู่").font = Font(bold=True)
    ws.cell(row=9, column=2, value="จำนวนเงิน").font = Font(bold=True)

    row = 10
    for cat, amt in sorted(cat_map.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=amt)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
